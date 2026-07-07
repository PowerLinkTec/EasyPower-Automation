"""
build_master.py — convert each EasyPower HTM report to an individual Excel file,
merge one-line diagram PDFs, and produce a combined PDF printout of all tables.

HTM reports:
  Every sc_*_det.htm / sc_*_sum.htm report becomes its own .xlsx file (e.g.
  sc_01_det.htm -> sc_01_det.xlsx) in the output folder.

PDF:
  One-line diagrams (SLDs): Every sc_*.pdf is concatenated into combined_sld.pdf
  (scenario order).
  Data tables: Every sc_*_*.xlsx is rendered into combined_report.pdf via
  reportlab.

Run standalone (interactive — asks where the reports are):
    python build_master.py
or quick:
    python build_master.py <reports_folder>
It also runs automatically at the end of easypower_batch_reports.py.

Needs: pip install pandas openpyxl lxml pypdf reportlab
"""

import re
import sys
from io import StringIO
from pathlib import Path

import numpy as np

# NumPy 2.x removed deprecated aliases like np.float, np.int, np.bool.
# Some libraries (e.g. openpyxl with older pandas) still reference them.
for _name, _val in (("float", float), ("int", int), ("bool", bool),
                     ("complex", complex)):
    if not hasattr(np, _name):
        setattr(np, _name, _val)

import pandas as pd

from pypdf import PdfWriter

# Some OpenSSL builds reject the 'usedforsecurity' keyword argument that
# pypdf and reportlab pass to md5().  Patch hashlib.md5 so it strips the
# keyword, allowing the call to fall through to the underlying C function.
import hashlib
_orig_md5 = hashlib.md5
def _md5_no_usedforsecurity(data=b"", **kwargs):
    kwargs.pop("usedforsecurity", None)
    return _orig_md5(data, **kwargs) if kwargs else _orig_md5(data)
hashlib.md5 = _md5_no_usedforsecurity


def _key(p):
    """Sort by scenario number.  Files that don't match sort at the end."""
    m = re.match(r"sc_(\d+)", p.stem)
    if m:
        return (int(m.group(1)), p.stem)
    return (float("inf"), p.stem)


def _report_title(stem):
    """sc_01_det -> 'Scenario 01 - Detailed', sc_01_sum -> 'Scenario 01 - Summary'."""
    m = re.match(r"sc_(\d+)_(det|sum)", stem)
    if not m:
        return stem
    num, rtype = m.group(1), m.group(2)
    return f"Scenario {num} - {'Detailed' if rtype == 'det' else 'Summary'}"


def _flatten_columns(df):
    """read_html turns multi-row table headers into MultiIndex columns, which
    pandas can't write to Excel with index=False. Join each column's levels into
    one name (dropping the empty/'Unnamed'/'nan' filler read_html inserts)."""
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [
            " ".join(s for s in (str(x).strip() for x in col)
                     if s and s.lower() != "nan" and not s.startswith("Unnamed")).strip()
            for col in df.columns
        ]
    return df


def _explode_multiline_rows(df):
    """Expand rows where cells contain \\n (from <br> in the HTM) into separate
    rows.  EasyPower sometimes puts multiple branch entries under one bus into a
    single HTML table row separated by <br> tags.  read_html reads those as a
    single cell with newlines; we split them so each entry gets its own Excel row,
    leaving columns that didn't have breaks blank on subsequent rows."""
    br_cols = [col for col in df.columns
               if df[col].astype(str).str.contains("\n", na=False).any()]
    if not br_cols:
        return df

    new_rows = []
    for _, row in df.iterrows():
        parts = {}
        n = 1
        for col in br_cols:
            val = str(row[col]) if pd.notna(row[col]) else ""
            lines = val.split("\n")
            parts[col] = lines
            n = max(n, len(lines))
        for i in range(n):
            nr = {}
            for col in df.columns:
                if col in br_cols:
                    lines = parts[col]
                    nr[col] = lines[i].strip() if i < len(lines) else ""
                else:
                    val = str(row[col]) if pd.notna(row[col]) else ""
                    nr[col] = val.strip() if i == 0 else ""
            new_rows.append(nr)
    return pd.DataFrame(new_rows)


def _preprocess_html(content):
    """EasyPower puts multiple entries in a single table row as separate
    <div> elements inside one <td>.  lxml's text_content() concatenates them
    without a separator, so we insert a unique delimiter before parsing."""
    return content.replace("</div><div", "</div>|||<div")


def _style_xlsx_sheet(ws):
    """Apply HTM-like styling to a worksheet after data has been written.

    Uses the same colour palette as the EasyPower-generated HTM reports:
    - Header row: #b3cae2 (blue-grey), bold, 10pt
    - Data rows: alternating #ffffff / #d9e4f0, 10pt
    - Thin borders, sensible column widths.
    """
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="B3CAE2", end_color="B3CAE2", fill_type="solid")
    ALT_FILL = PatternFill(start_color="D9E4F0", end_color="D9E4F0", fill_type="solid")

    HEADER_FONT = Font(bold=True, size=10)
    DATA_FONT = Font(size=10)

    THIN_BORDER = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    max_col = ws.max_column
    max_row = ws.max_row

    # Column widths: generous for text columns, tighter for numeric
    for col_idx in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Style header row (row 1) — matches HTM group/column header colour
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Style data rows — colour changes only when bus name (col A) changes
    bus_idx = -1
    for data_row in range(2, max_row + 1):
        a = ws.cell(row=data_row, column=1).value
        if a and str(a).strip():
            bus_idx += 1

        fill = ALT_FILL if bus_idx % 2 == 1 else None
        for col in range(1, max_col + 1):
            cell = ws.cell(row=data_row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill


def htm_to_excels(folder, out_dir=None):
    """Convert each sc_*_*.htm report into a separate .xlsx file.  Multiple
    <table> elements inside one HTM are stacked vertically on a single sheet."""
    if out_dir is None:
        out_dir = folder
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    htms = sorted(list(folder.glob("sc_*_*.htm")), key=_key)
    if not htms:
        print("No HTM reports found.")
        return

    converted = 0
    for f in htms:
        try:
            raw = f.read_text("utf-8")
            raw = _preprocess_html(raw)
            tables = pd.read_html(StringIO(raw))
        except Exception as e:
            print(f"skip {f.name}: {e}")
            continue

        out_path = out_dir / f"{f.stem}.xlsx"
        with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
            sheet, row = f.stem[:31], 0
            for t in tables:
                t = _flatten_columns(t)
                t = t.applymap(lambda v: v.replace("|||", "\n") if isinstance(v, str) else v)
                t = _explode_multiline_rows(t)
                t.to_excel(xl, sheet_name=sheet, index=False, startrow=row)
                row += len(t) + 2
            _style_xlsx_sheet(xl.sheets[sheet])
        converted += 1
        print(f"  {f.name} -> {out_path.name}")
    print(f"Converted {converted} report(s) to individual Excel files.")


def merge_pdfs(folder, out_pdf):
    """Merge one-line diagrams (SLDs): every sc_*.pdf -> one combined PDF."""
    pdfs = sorted(list(folder.glob("sc_*.pdf")), key=_key)
    if not pdfs:
        print("No PDF files found for merging.")
        return
    writer = PdfWriter()
    for f in pdfs:
        try:
            writer.append(str(f))
        except Exception as e:
            print(f"skip {f.name}: {e}")
    with open(out_pdf, "wb") as fh:
        writer.write(fh)
    print(f"Wrote {out_pdf} ({len(pdfs)} PDFs).")


def _calc_pdf_col_widths(data, avail_pt):
    """Distribute available page width proportionally across columns.

    Each column's share is based on its header text length (capped at 18 chars
    to prevent overly wide columns), scaled to exactly fill *avail_pt*.
    """
    if not data:
        return None
    header = data[0]
    char_width = 5.5
    raw = [max(30, min(110, len(str(h)) * char_width)) for h in header]
    total = sum(raw)
    return [w * avail_pt / total for w in raw]


def _word_wrap_cells(data, col_widths, font_size=8):
    """Insert ``\\n`` into cells whose text exceeds the column width.

    reportlab's Table draws cells with ``canvas.drawString`` which only
    respects explicit newlines.  This pre-processing step prevents long
    unbroken strings (e.g. branch-equipment names) from overflowing into
    adjacent cells.
    """
    char_width = font_size * 0.55
    for ri, row in enumerate(data):
        for ci, val in enumerate(row):
            text = str(val)
            if not text or ci >= len(col_widths):
                continue
            max_chars = int(col_widths[ci] / char_width) - 1
            if max_chars < 1 or len(text) <= max_chars:
                continue
            lines = []
            for i in range(0, len(text), max_chars):
                lines.append(text[i:i + max_chars])
            data[ri][ci] = "\n".join(lines)


def xlsx_to_combined_pdf(folder, out_pdf):
    """Convert every sc_*_*.xlsx into one combined PDF via reportlab.

    Each sheet from each workbook becomes a section in the PDF with a bold
    heading and a table rendered in 8pt Helvetica with grid lines,
    blue-grey header, per-bus-group alternating rows."""
    xlsx_files = sorted(list(folder.glob("sc_*_*.xlsx")), key=_key)
    if not xlsx_files:
        print("No Excel files found for PDF printout.")
        return

    from openpyxl import load_workbook
    from reportlab.lib.pagesizes import landscape, legal
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    doc = SimpleDocTemplate(
        str(out_pdf), pagesize=landscape(legal),
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=10*mm, bottomMargin=10*mm,
    )

    styles = getSampleStyleSheet()
    heading_style = styles["Heading2"]
    heading_style.textColor = colors.Color(0, 0, 0.6)
    elements = []

    for f in xlsx_files:
        wb = load_workbook(f, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            elements.append(
                Paragraph(f"<b>{_report_title(f.stem)}</b> &mdash; {sheet_name}", heading_style)
            )

            data = []
            for row in ws.iter_rows(values_only=True):
                data.append([str(v) if v is not None else "" for v in row])

            if data:
                col_widths = _calc_pdf_col_widths(data, doc.width)
                _word_wrap_cells(data, col_widths)

                # Build per-row BACKGROUND commands — colour changes on bus name
                bg_cmds = []
                bus_idx = -1
                last_bus = None
                for ri, row in enumerate(data[1:], start=1):
                    name = str(row[0]).strip() if row[0] else ""
                    if name and name != last_bus:
                        bus_idx += 1
                        last_bus = name
                    if bus_idx % 2 == 1:
                        bg_cmds.append(
                            ("BACKGROUND", (0, ri), (-1, ri),
                             colors.Color(0.85, 0.89, 0.94))
                        )

                t = Table(data, repeatRows=1, hAlign="LEFT", colWidths=col_widths)
                t.setStyle(
                    TableStyle([
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.6, 0.6, 0.6)),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.70, 0.79, 0.89)),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ] + bg_cmds)
                )
                elements.append(t)
            elements.append(PageBreak())
        wb.close()

    # Drop trailing blank page
    if elements and isinstance(elements[-1], PageBreak):
        elements.pop()

    if elements:
        doc.build(elements)
        print(f"Wrote {out_pdf} ({len(xlsx_files)} Excel files).")
    else:
        print("No data to write to PDF.")


def build(input_folder, output_folder=None):
    """Convert HTM reports to individual Excel files + merge PDFs + combined report."""
    in_dir = Path(input_folder)
    out_dir = Path(output_folder) if output_folder else in_dir / "ez_automation_output"
    out_dir.mkdir(parents=True, exist_ok=True)
    htm_to_excels(in_dir, out_dir)
    merge_pdfs(in_dir, out_dir / "combined_sld.pdf")
    xlsx_to_combined_pdf(out_dir, out_dir / "combined_report.pdf")


def banner():
    """Styled title block (matches easypower_batch_reports.py). Duplicated here so
    this script doesn't have to import the pywinauto-heavy main module."""
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    lines = [
        "EasyPower Reporting Automation",
        "Copyright © Power-Link Technologies 2026",
        "",
        "Reach out to cooper@powerlinktec.com",
        "for troubleshooting and for reporting bugs",
    ]
    w = max(len(s) for s in lines) + 6
    print("\n╔" + "═" * w + "╗")
    for s in lines:
        print("║" + s.center(w) + "║")
    print("╚" + "═" * w + "╝\n")


if __name__ == "__main__":
    if len(sys.argv) > 1:                                # quick: folder on the command line
        build(sys.argv[1])
    else:
        banner()
        while True:
            src = input("Folder containing the report files: ").strip().strip('"')
            if Path(src).is_dir():
                break
            print("   That folder doesn't exist — try again.")
        default_out = str(Path(src) / "ez_automation_output")
        dst = input(f"Output folder (for the .xlsx + .pdf files) [{default_out}]: ").strip().strip('"')
        build(src, dst or None)
